import openpyxl
import os
import PySimpleGUI as sg

def findrow(lrn):
    #row to find
    global subjectrow
    global lrnrow
    subjectrow = 0
    lrnrow = 0
    for i in range(1,srcmaxrow+1):
        cell = srcsheet.cell(i, 3)
        if not(cell.value == None):
            if cell.value == 'LRN' and (subjectrow == 0):
                subjectrow = i-1
            if "".join(cell.value.split()) == "".join(lrn.split()):
                lrnrow = i 
                return True
    return False
            
def findgrade():
    global subjectgrade
    subjectgrade = {}
    for i in range(1, srcmaxcolumn+1):
        if i >= 4:
            cell = srcsheet.cell(subjectrow,i)
            if cell.value != None:
                grade = []
                for j in range(2):
                        if type(srcsheet.cell(lrnrow,i+j).value) == int:
                            grade.append(srcsheet.cell(lrnrow,i+j).value)
                        else:
                            return
                if('Physical Education' in " ".join(cell.value.split())):
                    #change numbering for PE
                    subjectgrade['Physical Education and Health'] = grade
                else: 
                    subjectgrade[" ".join(cell.value.split())] = grade

def findsubjectcategory(subj):
    for cat in subjcatrow:
        print('hey')
        for i in range(subjcatrow[cat][0],subjcatrow[cat][1]+1):
            if('Physical Education' in subj and 'Physical Education' in annexsheet.cell(i , 3).value):
                return cat
            if(annexsheet.cell(i , 3).value == subj):
                return cat

def inputdata(name):
    global gradelvl
    global semester
    global dirsheet
    if gradelvl == '12':
        dirsheet = dirworkbook['BACK']
    elif gradelvl == '11':
        dirsheet = dirworkbook['FRONT']
    global annexsheet
    annexsheet = dirworkbook['ANNEX']

    inputrow = findf137Row()
    print(inputrow)
    #inputing data
    currow = inputrow
    for subj in subjectgrade:
        #category
        dirsheet.cell(currow,1).value = findsubjectcategory(subj)
        #subject name
        dirsheet.cell(currow,9).value = subj
        #1st/3rd grade
        dirsheet.cell(currow,46).value = subjectgrade[subj][0]
        #2nd/4th grade
        dirsheet.cell(currow,51).value = subjectgrade[subj][1]
        currow = currow+1
    
    #save the excel file
    dirworkbook.save(savepath + name)

def resultsform(resultdata):
    global stringdata
    stringdata = ""
    for i in resultdata:
        stringdata += (i + ' : ' + resultdata[i] + '\n')
    return stringdata

def inputchecker():
    for i in values:
        if values[i] == '':
            return True
    return False

def main():
    global directorypath
    global dirworkbook
    loading = 1
    for i in os.listdir(directories):
        print('processing: ',i)
        sg.one_line_progress_meter('One Line Meter: ', loading ,len(os.listdir(directories)))
        print(loading)
        loading = loading + 1
        if i.split('.')[-1] == 'xlsx': 
            findlrn = i.split('-')[0]
            #find where are the SUBJECT rows of data on source
            if (findrow(findlrn)):
                directorypath = directories + i
                dirworkbook = openpyxl.load_workbook(directorypath)
                findgrade()
                #transfer subject and data to directory
                inputdata(i)
                print("succes: ", findlrn)
                result[i]='Success'
            else:
                print("cant find lrn: ", findlrn)
                result[i]='Cant find LRN'
        else:
            print("not an xlsx file:", i)
            result[i]='Not an excel file'

def findf137Row():
    sheet = dirsheet
    sheetmax = sheet.max_row
    counter = 0
    for i in range(1, sheetmax+1):
        if sheet.cell(i,1).value != None:
            if 'Indicate' in sheet.cell(i,1).value:
                counter += 1
                if semester == '1' and counter == 1:
                    return i+4
                if semester == '2' and counter == 3:
                    return i+4
#dictionaries
"""
f137row = {
    #rows of data [gradelvl][semester]:[startingrow on excel]
    #row of g11 1st sem
    '111':31,
    #row of g11 2nd sem
    '112':74,
    #row of g12 1st sem
    '121':11,
    #row of g12 2nd sem
    '122':54
}
"""
subjcatrow = {
    #[category]:[startingrow on ANNEX,eding row on ANNEX]
    'Core': [7,27],
    'Applied': [30,36],
    'Specialized': [39,52],
    'Other_Subjects': [54, 56]
}

result = {}



#GUI
#Theme
sg.theme('DarkAmber')
#Layout
layout = [
    [sg.Text('SOG file path:')],
    [sg.Input(readonly=True,text_color='black'),sg.FileBrowse(file_types=(('Excel Files','*.xlsx'),),key='sourceinput')],
    [sg.Text('Directory folder of F137:')],
    [sg.Input(readonly=True,text_color='black'),sg.FolderBrowse(key='directoryinput',tooltip='Specify the F137 folder\n*please follow name format')],
    [sg.Text('Save path:')],
    [sg.Input(readonly=True,text_color='black'),sg.FolderBrowse(key='savepath')],
    [sg.Push(), sg.Text('Grade'), sg.OptionMenu(values=('12','11'), size = 11, key='Grade', default_value= '11'),
    sg.Text('Semester'),sg.OptionMenu(values=('1','2'), size = 11, key='Semester', default_value='1'), sg.Push()],
    [sg.Push(),sg.Submit('Start',size=10),sg.Button('Instructions',size=10), sg.Exit(size=10),sg.Push()]
]


#Window
window = sg.Window('SOG to F137(BETA)', layout)

sg.popup(
        #opening message
        '-Summary of grades to form 137-\n\n'+
        'THIS PROGRAM IS STILL IN BETA.\n'+
        'The program will close if an error is occured.\n'+
        'Please follow the format before loading the file\n\n'+
        'For questions please contact the developer:\n'+
        'rattapuroc@gmail.com',
        no_titlebar=True
    )

# GUI Event loop 
while True:
    
    event, values = window.read()
    print(event, values)
    
    if event == 'Start':
            
        if (inputchecker()):
            sg.popup('Please fill out all fields',no_titlebar=True, background_color='white',text_color='black')
        elif (values['directoryinput'] == values['savepath']):
            sg.popup('F137 folder and savepath must not be the same',no_titlebar=True, background_color='white',text_color='black')
        else:
            #load source file
            sourcepath = values['sourceinput'].replace(os.sep,'/')
            srcworkbook = openpyxl.load_workbook(sourcepath)
            srcsheet = srcworkbook.active

            #get max column and max row of source
            srcmaxrow = srcsheet.max_row
            srcmaxcolumn = srcsheet.max_column

            #load directory folder
            dirspath = values['directoryinput'] + '/'
            directories = dirspath.replace(os.sep, '/')

            #load grade and semester
            gradelvl = values['Grade']
            semester = values['Semester']

            #load save directory folder
            savepath = values['savepath'] + '/'

            sg.popup_no_buttons('starting', no_titlebar=True, auto_close_duration= 2, auto_close= True)
            main()
            print(result)
            sg.popup('Program summary' ,resultsform(result))

    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    if event == 'Instructions':
        sg.popup_non_blocking(
            '1. Browse the file of summary of grades\n'+
            '   *Follow the SOG excel format (ask developer)\n'+
            '2. Browse folder where Form 137 are located\n'+
            '   *On Form137 file name put (LRN) first then (-)\n    *Use the base format of Form137 excel\n'+
            '3. Browse save folder of the excel files\n'+
            '   *Do not make step 2 and 3 same!\n'+
            '4. Select Grade and Semester\n\nRaw values:\n\n'+
            'Subject category row on ANNEX sheet(F137):\ncategory:(start row, endrow)\n\n'+ str(subjcatrow)
            )
